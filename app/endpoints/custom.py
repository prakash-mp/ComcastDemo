from typing import List
from io import BytesIO

from fastapi import Depends, status, File, UploadFile
from fastapi.encoders import jsonable_encoder
from fastapi.responses import JSONResponse
from fastapi.routing import APIRouter
from sqlalchemy.orm import Session
from typing_extensions import Annotated
from openpyxl import load_workbook

from app import models, schemas, log
from app.dependencies import deps

router = APIRouter(tags=["Custom Simulator"])


@router.post("/custom/hubs")
async def create_custom_hub(
    db: Annotated[Session, Depends(deps.get_db_session)], file: UploadFile = File(...)
):
    try:
        # Read the contents of the file into memory
        contents = await file.read()
        excel_file = BytesIO(contents)

        # Use openpyxl to load the workbook
        wb = load_workbook(excel_file, data_only=True)
        sheet = wb.active  # Get the first sheet

        # Get headers from the first row (you can modify this if you need custom header handling)
        headers = [cell.value for cell in sheet[1]]  # First row for headers

        # Validate the necessary headers
        required_headers = [
            "hub_id",
            "partner",
            "hubCode",
            "hubName",
            "hubType",
            "primaryHubId",
            "addr1",
            "addr2",
            "city",
            "BuhmId",
            "countryCode",
            "locality",
            "serviceStatus",
            "state",
            "zipCode",
            "timezone",
            "type",
            "coordinates",
            "createdBy",
            "modifiedBy",
        ]

        if not all(header in headers for header in required_headers):
            log.info("Excel file is missing required headers.")
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content={
                    "code": 400,
                    "status": "Bad Request",
                    "message": "Excel file is missing required headers.",
                },
            )

        db_objs = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Each row is a tuple, where each value corresponds to a cell in that row
            row_data = dict(zip(headers, row))
            hub_id = row_data["hub_id"]

            db_obj = (
                db.query(models.Comcast).filter(models.Comcast.hub_id == hub_id).first()
            )
            if db_obj:
                log.info(f"Record already created for hub_id {hub_id}")
                return JSONResponse(
                    status_code=status.HTTP_409_CONFLICT,
                    content={"code": 409, "status": "Conflict", "message": "Error"},
                )

            hub_data = {
                "hub_id": hub_id,
                "partner": row_data["partner"],
                "hub": {
                    "hubCode": row_data["hubCode"],
                    "hubName": row_data["hubName"],
                    "hubType": row_data["hubType"],
                    "primaryHubId": row_data["primaryHubId"],
                    "addr1": row_data["addr1"],
                    "addr2": row_data["addr2"],
                    "city": row_data["city"],
                    "BuhmId": row_data["BuhmId"],
                    "countryCode": row_data["countryCode"],
                    "locality": row_data["locality"],
                    "serviceStatus": row_data["serviceStatus"],
                    "state": row_data["state"],
                    "zipCode": row_data["zipCode"],
                    "timezone": row_data["timezone"],
                    "location": {
                        "type": row_data["type"],
                        "coordinates": row_data["coordinates"],
                    },
                },
                "createdBy": row_data["createdBy"],
                "modifiedBy": row_data["modifiedBy"],
            }

            data_in = schemas.ComcastCreate(**hub_data)
            db_obj = models.Comcast.from_schema(data_in)
            db.add(db_obj)
            db_objs.append(db_obj)

        db.commit()
        created_hub_ids = [db_obj.hub_id for db_obj in db_objs]
        log.info(
            f"All custom record created for uploaded excel for hub_ids {created_hub_ids}"
        )
        return JSONResponse(
            status_code=status.HTTP_201_CREATED,
            content={
                "code": 201,
                "status": "OK",
                "message": [jsonable_encoder(db_obj.to_schema()) for db_obj in db_objs],
            },
        )

    except Exception as exc:
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={
                "code": 500,
                "status": "Internal Server Error",
                "message": f"Error processing file: {str(exc)}",
            },
        )


@router.get("/custom/hubs/{hub_id}")
def get_custom_hub(db: Annotated[Session, Depends(deps.get_db_session)], hub_id: str):
    db_obj = db.query(models.Comcast).filter(models.Comcast.hub_id == hub_id).first()
    if not db_obj:
        log.info(f"No record found for hub_id {hub_id}")
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={
                "code": 404,
                "status": "Not found",
                "message": "Error",
            },
        )

    return JSONResponse(
        status_code=status.HTTP_200_OK,
        content={
            "code": 200,
            "status": "OK",
            "message": jsonable_encoder(db_obj.to_schema()),
        },
    )


@router.get("/custom/hubs")
def get_all_custom_hub(db: Annotated[Session, Depends(deps.get_db_session)]):
    db_objs: List[models.Comcast] = db.query(models.Comcast).all()

    return JSONResponse(
        status_code=status.HTTP_200_OK,
        content={
            "code": 200,
            "status": "OK",
            "message": [jsonable_encoder(db_obj.to_schema()) for db_obj in db_objs],
        },
    )


@router.patch("/custom/hubs/{hub_id}")
def update_custom_hub(
    db: Annotated[Session, Depends(deps.get_db_session)],
    hub_id: str,
    data_in: schemas.ComcastUpdate,
):
    db_obj = db.query(models.Comcast).filter(models.Comcast.hub_id == hub_id).first()
    if not db_obj:
        log.info(f"No record found for hub_id {hub_id}")
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={
                "code": 404,
                "status": "Not found",
                "message": "Error",
            },
        )

    # tbi


@router.delete("/custom/hubs/{hub_id}")
def delete_custom_hub(
    db: Annotated[Session, Depends(deps.get_db_session)], hub_id: str
):
    db_obj = db.query(models.Comcast).filter(models.Comcast.hub_id == hub_id).first()
    if not db_obj:
        log.info(f"No record found for hub_id {hub_id}")
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={
                "code": 404,
                "status": "Not found",
                "message": "Error",
            },
        )

    log.info(f"Record deleted for hub_id {hub_id}")
    db.delete(db_obj)
    db.commit()

    return status.HTTP_204_NO_CONTENT
